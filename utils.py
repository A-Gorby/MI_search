import pandas as pd
# # print(pd.__version__)
import numpy as np
import os, sys, glob
import humanize
import re
import xlrd

import json
import itertools
from itertools import zip_longest
#from urllib.request import urlopen
#import requests, xmltodict
import time, datetime
import math
from pprint import pprint
import gc
from tqdm import tqdm
tqdm.pandas()
import pickle

import logging
import zipfile
import warnings
import argparse

import warnings
warnings.filterwarnings("ignore")
# import numba
# import numexpr as ne
# numba.set_num_threads(numba.get_num_threads())
import requests
from urllib.parse import urlencode

from fuzzywuzzy import fuzz
from fuzzywuzzy import process

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.utils import units
from openpyxl.styles import Border, Side, PatternFill, GradientFill, Alignment

from sentence_transformers import SentenceTransformer, util
import ipywidgets as widgets
from IPython.display import display
from ipywidgets import Layout, Box, Label

class Logger():
    def __init__(self, name = 'Fuzzy Lookup',
                 strfmt = '[%(asctime)s] [%(levelname)s] > %(message)s', # strfmt = '[%(asctime)s] [%(name)s] [%(levelname)s] > %(message)s'
                 level = logging.INFO,
                 datefmt = '%H:%M:%S', # '%Y-%m-%d %H:%M:%S'
                #  datefmt = '%H:%M:%S %p %Z',

                 ):
        self.name = name
        self.strfmt = strfmt
        self.level = level
        self.datefmt = datefmt
        self.logger = logging.getLogger(name)
        self.logger.setLevel(self.level) #logging.INFO)
        self.offset = datetime.timezone(datetime.timedelta(hours=3))
        # create console handler and set level to debug
        self.ch = logging.StreamHandler()
        self.ch.setLevel(self.level)
        # create formatter
        self.strfmt = strfmt # '[%(asctime)s] [%(levelname)s] > %(message)s'
        self.datefmt = datefmt # '%H:%M:%S'
        # создаем форматтер
        self.formatter = logging.Formatter(fmt=strfmt, datefmt=datefmt)
        self.formatter.converter = lambda *args: datetime.datetime.now(self.offset).timetuple()
        self.ch.setFormatter(self.formatter)
        # add ch to logger
        self.logger.addHandler(self.ch)

logger = Logger().logger
logger.propagate = False

def save_df_to_excel(df, path_to_save, fn_main, columns = None, b=0, e=None, index=False):
    offset = datetime.timezone(datetime.timedelta(hours=3))
    dt = datetime.datetime.now(offset)
    str_date = dt.strftime("%Y_%m_%d_%H%M")
    fn = fn_main + '_' + str_date + '.xlsx'
    logger.info(fn + ' save - start ...')
    #if e is None or (e <0):
    #    e = df.shape[0]
    if columns is None:
        if e is None or (e <0):
            df.to_excel(os.path.join(path_to_save, fn), index = index)
        else: 
            df[b:e].to_excel(os.path.join(path_to_save, fn), index = index)
    else:
        if e is None or (e <0):
            df.to_excel(os.path.join(path_to_save, fn), index = index, columns = columns)
        else:
            df[b:e].to_excel(os.path.join(path_to_save, fn), index = index, columns = columns)
    logger.info(fn + ' saved to ' + path_to_save)
    hfs = get_humanize_filesize(path_to_save, fn)
    logger.info("Size: " + str(hfs))
    return fn
def get_humanize_filesize(path, fn):
    human_file_size = None
    try:
        fn_full = os.path.join(path, fn)
    except Exception as err:
        print(err)
        return human_file_size
    if os.path.exists(fn_full):
        file_size = os.path.os.path.getsize(fn_full)
        human_file_size = humanize.naturalsize(file_size)
    return human_file_size
    
def restore_df_from_pickle(path_files, fn_pickle):

    if fn_pickle is None:
        logger.error('Restore pickle from ' + path_files + ' failed!')
        sys.exit(2)
    if os.path.exists(os.path.join(path_files, fn_pickle)):
        df = pd.read_pickle(os.path.join(path_files, fn_pickle))
        # logger.info('Restore ' + re.sub(path_files, '', fn_pickle_с) + ' done!')
        logger.info('Restore ' + fn_pickle + ' done!')
        logger.info('Shape: ' + str(df.shape))
    else:
        # logger.error('Restore ' + re.sub(path_files, '', fn_pickle_с) + ' from ' + path_files + ' failed!')
        logger.error('Restore ' + fn_pickle + ' from ' + path_files + ' failed!')
    return df    

# from utils import unzip_file
def unzip_file(path_source, fn_zip, work_path):
    logger.info('Unzip ' + fn_zip + ' start...')

    try:
        with zipfile.ZipFile(path_source + fn_zip, 'r') as zip_ref:
            fn_list = zip_ref.namelist()
            zip_ref.extractall(work_path)
        logger.info('Unzip ' + fn_zip + ' done!')
        return fn_list[0]
    except Exception as err:
        logger.error('Unzip error: ' + str(err))
        sys.exit(2)


def save_to_excel(df_total, total_sheet_names, save_path, fn):
    # fn = model + '.xlsx'
    offset = datetime.timezone(datetime.timedelta(hours=3))
    dt = datetime.datetime.now(offset)
    str_date = dt.strftime("%Y_%m_%d_%H%M")
    fn_date = fn.replace('.xlsx','')  + '_' + str_date + '.xlsx'
    
    # with pd.ExcelWriter(os.path.join(path_tkbd_processed, fn_date )) as writer:  
    with pd.ExcelWriter(os.path.join(save_path, fn_date )) as writer:  
        
        for i, df in enumerate(df_total):
            df.to_excel(writer, sheet_name = total_sheet_names[i], index=False)
    return fn_date    

def get_humanize_filesize(path, fn):
    human_file_size = None
    try:
        fn_full = os.path.join(path, fn)
    except Exception as err:
        print(err)
        return human_file_size
    if os.path.exists(fn_full):
        file_size = os.path.os.path.getsize(fn_full)
        human_file_size = humanize.naturalsize(file_size)
    return human_file_size
    
def np_unique_nan(lst: np.array, debug = False)->np.array: # a la version 2.4
    lst_unique = None
    if lst is None or (((type(lst)==float) or (type(lst)==np.float64)) and np.isnan(lst)):
        # if debug: print('np_unique_nan:','lst is None or (((type(lst)==float) or (type(lst)==np.float64)) and math.isnan(lst))')
        lst_unique = lst
    else:
        data_types_set = list(set([type(i) for i in lst]))
        if debug: print('np_unique_nan:', 'lst:', lst, 'data_types_set:', data_types_set)
        if ((type(lst)==list) or (type(lst)==np.ndarray)):
            if debug: print('np_unique_nan:','if ((type(lst)==list) or (type(lst)==np.ndarray)):')
            if len(data_types_set) > 1: # несколько типов данных
                if list not in data_types_set and dict not in data_types_set \
                      and tuple not in data_types_set and type(None) not in data_types_set\
                      and np.ndarray not in data_types_set: # upd 17/02/2023
                    lst_unique = np.array(list(set(lst)), dtype=object)
                else:
                    lst_unique = lst
            elif len(data_types_set) == 1:
                if debug: print("np_unique_nan: elif len(data_types_set) == 1:")
                if list in data_types_set:
                    lst_unique = np.unique(np.array(lst, dtype=object))
                elif  np.ndarray in data_types_set:
                    # print('elif  np.ndarray in data_types_set :')
                    # lst_unique = np.unique(lst.astype(object))
                    lst_unique = np_unique_nan(lst_unique)
                    lst_unique = np.asarray(lst, dtype = object)
                    # lst_unique = np.unique(lst_unique)
                elif type(None) in data_types_set:
                    # lst_unique = np.array(list(set(lst)))
                    lst_unique = np.array(list(set(list(lst))))
                elif dict in  data_types_set:
                    lst_unique = lst
                    # np.unique(lst)
                elif type(lst) == np.ndarray:
                    if debug: print("np_unique_nan: type(lst) == np.ndarray")
                    if (lst.dtype.kind == 'f') or  (lst.dtype == np.float64) or  (float in data_types_set):
                        if debug: print("np_unique_nan: (lst.dtype.kind == 'f')")
                        lst_unique = np.unique(lst.astype(float))
                        # if debug: print("np_unique_nan: lst_unique predfinal:", lst_unique)
                        # lst_unique = np.array(list(set(list(lst))))
                        # if debug: print("np_unique_nan: lst_unique predfinal v2:", lst_unique)
                        # if np.isnan(lst).all():
                        #     lst_unique = np.nan
                        #     if debug: print("np_unique_nan: lst_unique predfinal v3:", lst_unique)
                    elif (lst.dtype.kind == 'S') :
                        if debug: print("np_unique_nan: lst.dtype == string")
                        lst_unique = np.array(list(set(list(lst))))
                        if debug: print(f"np_unique_nan: lst_unique 0: {lst_unique}")
                    elif lst.dtype == object:
                        if debug: print("np_unique_nan: lst.dtype == object")
                        if (type(lst[0])==str) or (type(lst[0])==np.str_) :
                            try:
                                lst_unique = np.unique(lst)
                            except Exception as err:
                                lst_unique = np.array(list(set(list(lst))))
                        else:
                            lst_unique = np.array(list(set(list(lst))))
                        if debug: print(f"np_unique_nan: lst_unique 0: {lst_unique}")
                    else:
                        if debug: print("np_unique_nan: else 0")
                        lst_unique = np.unique(lst)
                else:
                    if debug: print('np_unique_nan:','else i...')
                    lst_unique = np.array(list(set(lst)))
                    
            elif len(data_types_set) == 0:
                lst_unique = None
            else:
                # print('else')
                lst_unique = np.array(list(set(lst)))
        else: # другой тип данных
            if debug: print('np_unique_nan:','другой тип данных')
            # lst_unique = np.unique(np.array(list(set(lst)),dtype=object))
            # lst_unique = np.unique(np.array(list(set(lst)))) # Исходим из того что все елеменыт спсика одного типа
            lst_unique = lst
    if type(lst_unique) == np.ndarray:
        if debug: print('np_unique_nan: final: ', "if type(lst_unique) == np.ndarray")
        if lst_unique.shape[0]==1: 
            if debug: print('np_unique_nan: final: ', "lst_unique.shape[0]==1")
            lst_unique = lst_unique[0]
            if debug: print(f"np_unique_nan: final after: lst_unique: {lst_unique}")
            if (type(lst_unique) == np.ndarray) and (lst_unique.shape[0]==1):  # двойная вложенность
                if debug: print('np_unique_nan: final: ', 'one more', "lst_unique.shape[0]==1")
                lst_unique = lst_unique[0]
        elif lst_unique.shape[0]==0: lst_unique = None
    if debug: print(f"np_unique_nan: return: lst_unique: {lst_unique}")
    if debug: print(f"np_unique_nan: return: type(lst_unique): {type(lst_unique)}")
    return lst_unique

import re, regex
def search_product_options_by_str_list(s_lst):
    '''
    sorce: ["Проволка ортопедическая"] or ["Проволка", "ортопедическая"] or ["Спица Киршнера", "1,8", "мм"]
    target: "Проволка костная ортопедическая" or в описании продукта (варианты исполнения) 
         --- Спица Киршнера с троакарным концом. ...., - диаметр 1,6 мм, длина 150-310 мм, - диаметр 1,8 мм, длина 150-310 мм,
    '''
    global df_mi_org_gos
    code_gos, name_gos, product_options_gos, code_national, name_national = None, None, None, None, None
    if (type(s_lst)==list) or (type(s_lst)==np.ndarray) and (len(s_lst)>0) and ((type(s_lst[0])==str) or (type(s_lst[0])==np.str_)):
        query_str_01, query_str_02 = '', ''
        # if len(s_lst)>1:
        for i, el in enumerate(s_lst):
            if len(query_str_01) > 0:
                # query_str_01 += f" & product_options.str.contains('{s_lst[i]}', case=False)" 
                query_str_01 += f" & product_options.str.contains('{s_lst[i]}', regex=True, flags=re.I)" 
            else: 
                # query_str_01 += f"product_options.notnull() & product_options.str.contains('{s_lst[i]}', case=False)"
                # query_str_01 += f"product_options.notnull() & product_options.str.contains('{s_lst[i]}', regex=True, flags=re.I)"
                query_str_01 += f"product_options.notnull() & product_options.str.contains('{s_lst[i]}', regex=True, flags=@re.I)"
        # else:
        #     query_str_01 = "product_options == @s_lst[0]"
        lst_01 = df_mi_org_gos.query(query_str_01, engine='python')[['kind', 'name_clean', 'product_options']].values # , 'product_options'
        if len(lst_01) >0:
            code_gos = np_unique_nan(lst_01[:,0])
            name_gos = np_unique_nan(lst_01[:,1])
            product_options_gos = lst_01[:2]
        return code_gos, name_gos, product_options_gos
        
        
    else:
        return code_gos, name_gos, product_options_gos    

def load_check_dictionaries_for_fuzzy_search(path_supp_dicts,
      fn_df_mi_org_gos ='df_mi_org_gos_release_20230129_2023_02_14_1759.pickle',
      fn_df_mi_org_gos_prod_options ='df_mi_org_gos_prod_options_release_20230201_2023_02_14_1835.pickle',
      fn_df_mi_national = 'df_mi_national_release_20230201_2023_02_06_1013.pickle',
      fn_dict_embedding_gos_multy = 'dict_embedding_gos_multy.pickle', 
      fn_dict_embedding_gos_prod_options_multy = 'dict_embedding_gos_prod_options_multy.pickle',
      fn_dict_embedding_national_multy = 'dict_embedding_national_multy.pickle',
      fn_dict_lst_gos_prod_options ='dict_lst_gos_prod_options.pickle',
    ):
    # global df_services_MGFOMS, df_services_804n, df_RM, df_MNN, df_mi_org_gos, df_mi_national, df_mi_org_gos_prod_options
    
    df_mi_org_gos, df_mi_org_gos_prod_options, df_mi_national = None, None, None
    dict_embedding_gos_multy, dict_embedding_gos_prod_options_multy, dict_embedding_national_multy = None, None, None
    
    # fn_df_mi_org_gos = 'df_mi_org_gos_release_20230129_2023_02_07_1331.pickle'
    # fn_df_mi_national = 'df_mi_national_release_20230201_2023_02_06_1013.pickle'
    # fn_df_mi_org_gos ='df_mi_org_gos_release_20230129_2023_02_14_1759.pickle'
    # fn_df_mi_org_gos_prod_options ='df_mi_org_gos_prod_options_release_20230201_2023_02_14_1835.pickle'
    df_mi_org_gos = restore_df_from_pickle(path_supp_dicts, fn_df_mi_org_gos)
    df_mi_org_gos_prod_options = restore_df_from_pickle(path_supp_dicts, fn_df_mi_org_gos_prod_options)
    df_mi_national = restore_df_from_pickle(path_supp_dicts, fn_df_mi_national)
    dict_embedding_gos_multy = restore_df_from_pickle(path_supp_dicts, fn_dict_embedding_gos_multy )
    dict_embedding_gos_prod_options_multy = restore_df_from_pickle(path_supp_dicts, fn_dict_embedding_gos_prod_options_multy )
    dict_embedding_national_multy = restore_df_from_pickle(path_supp_dicts, fn_dict_embedding_national_multy )
    dict_lst_gos_prod_options = restore_df_from_pickle(path_supp_dicts, fn_dict_lst_gos_prod_options)
    
    return df_mi_org_gos, df_mi_org_gos_prod_options, df_mi_national, \
          dict_embedding_gos_multy, dict_embedding_gos_prod_options_multy, dict_embedding_national_multy, dict_lst_gos_prod_options


def upload_files_for_fuzzy_search(supp_dict_dir = '/content/data/supp_dict', links = {
    'df_mi_national': {'fn': 'df_mi_national_release_20230201_2023_02_06_1013.zip', 'ya_link': 'https://disk.yandex.ru/d/pfgyT_zmcYrHBw' },
    'df_mi_org_gos' : {'fn': 'df_mi_org_gos_release_20230129_2023_02_14_1759.zip', 'ya_link': 'https://disk.yandex.ru/d/xYolPYsHiSFEWA' },
    'df_mi_org_gos_prod_options': {'fn': 'df_mi_org_gos_prod_options_release_20230201_2023_02_14_1835.zip', 'ya_link': 'https://disk.yandex.ru/d/fnBfPpB8L-mJaw' },
    'dict_embedding_gos_multy' :{'fn': 'dict_embedding_gos_multy.pickle', 'ya_link': 'https://disk.yandex.ru/d/mArd7T-od6NcaQ'},
    'dict_embedding_gos_prod_options_multy': {'fn': 'dict_embedding_gos_prod_options_multy.pickle', 'ya_link': 'https://disk.yandex.ru/d/c2PdgI4JCbnWaA'},
    'dict_embedding_national_multy' : {'fn': 'dict_embedding_national_multy.pickle', 'ya_link': 'https://disk.yandex.ru/d/2qio4quws5IcUQ'},
    'dict_lst_gos_prod_options' : {'fn': 'dict_lst_gos_prod_options.pickle',  'ya_link': 'https://disk.yandex.ru/d/QBhdbktm2_25Ew'},
}):
    base_url = 'https://cloud-api.yandex.net/v1/disk/public/resources/download?'
    # public_key = link #'https://yadi.sk/d/UJ8VMK2Y6bJH7A'  # Сюда вписываете вашу ссылку

    # Получаем загрузочную ссылку
    for link in tqdm(links.values()):
        final_url = base_url + urlencode(dict(public_key=link['ya_link']))
        
        response = requests.get(final_url)
        download_url = response.json()['href']

        # Загружаем файл и сохраняем его
        download_response = requests.get(download_url)
        # with open('downloaded_file.txt', 'wb') as f:   # Здесь укажите нужный путь к файлу
        with open(os.path.join(supp_dict_dir, link['fn']), 'wb') as f:   # Здесь укажите нужный путь к файлу
            f.write(download_response.content)
            logger.info(f"File '{link['fn']}' uploaded!")
            if link['fn'].split('.')[-1] == 'zip':
                fn_unzip = unzip_file(os.path.join(supp_dict_dir, link['fn']), '', supp_dict_dir)
                logger.info(f"File '{fn_unzip}' upzipped!")    

def upload_check_dictionaries(supp_dict_dir, data_links):
    upload_files_for_fuzzy_search(supp_dict_dir, links = data_links)
    df_mi_org_gos, df_mi_org_gos_prod_options, df_mi_national, \
    dict_embedding_gos_multy, dict_embedding_gos_prod_options_multy, dict_embedding_national_multy, dict_lst_gos_prod_options = \
      load_check_dictionaries_for_fuzzy_search( supp_dict_dir,
            fn_df_mi_org_gos = data_links['df_mi_org_gos']['fn'],
            fn_df_mi_org_gos_prod_options = data_links['df_mi_org_gos_prod_options']['fn'],
            fn_df_mi_national = data_links['df_mi_national']['fn'],
            fn_dict_embedding_gos_multy = data_links['dict_embedding_gos_multy']['fn'],
            fn_dict_embedding_gos_prod_options_multy = data_links['dict_embedding_gos_prod_options_multy']['fn'],
            fn_dict_embedding_national_multy = data_links['dict_embedding_national_multy']['fn'],
            fn_dict_lst_gos_prod_options =data_links['dict_lst_gos_prod_options']['fn'],
    )
    return df_mi_org_gos, df_mi_org_gos_prod_options, df_mi_national, \
    dict_embedding_gos_multy, dict_embedding_gos_prod_options_multy, dict_embedding_national_multy, dict_lst_gos_prod_options

def test_big_dictionaries(df_mi_org_gos, df_mi_org_gos_prod_options, df_mi_national, 
      dict_embedding_gos_multy, dict_embedding_gos_prod_options_multy, dict_embedding_national_multy, dict_lst_gos_prod_options):
    if df_mi_org_gos is None or df_mi_org_gos_prod_options is None or df_mi_national is None or  \
        dict_embedding_gos_multy is None or dict_embedding_gos_prod_options_multy is None or dict_embedding_national_multy is None \
        or dict_lst_gos_prod_options is None:
        logger.error(f"Большие справочники не загружены")
        test_ok = False
    else: test_ok = True
    return test_ok

def test_inputs(data_source_dir, 
              fn_check_file, sheet_name_check, col_name_check,
              fn_dict_file, sheet_name_dict, col_names_dict, by_big_dict
              ):
    # global df_mi_org_gos, df_mi_org_gos_prod_options, df_mi_national, dict_embedding_gos_multy, dict_embedding_gos_prod_options_multy, dict_embedding_national_multy, dict_lst_gos_prod_options
    test_ok = True
    if not os.path.exists (data_source_dir):
        test_ok = False
        # logger.error(f"Path for source data '{data_source_dir}' not exists")
        logger.error(f"Директория для входных данных '{data_source_dir}' не существует")
    elif not os.path.isdir(data_source_dir):
        test_ok = False
        #logger.error(f"'{data_source_dir}' is not a dir")
        logger.error(f"Директория для входных данных '{data_source_dir}' - не директория")
    else:
        if fn_check_file is None or not os.path.exists(os.path.join(data_source_dir, fn_check_file)):
            test_ok = False
            # logger.error(f"Check file '{fn_check_file}' not found")
            logger.error(f"Проверяемый файл '{fn_check_file}' не найден")
        
        if fn_dict_file is None and by_big_dict:
            pass
        elif not os.path.exists(os.path.join(data_source_dir, fn_dict_file)):
            test_ok = False
            # logger.error(f"Dictionary file '{fn_dict_file}' not found")
            logger.error(f"Файл справочника '{fn_dict_file}' не найден")
    return test_ok

def read_check_file(path_check_file, fn_check_file, sheet_name, col_name= 'Наименование МИ/РМ', debug=False):
    df = None
    read_ok = True
    try:
        if sheet_name is None:
            wb = load_workbook(os.path.join(path_check_file, fn_check_file))
            for sh_n in wb.get_sheet_names():
                if 'тест' in sh_n.lower():
                    sheet_name = sh_n
            if sheet_name is None: sheet_name = wb.get_sheet_names()[0]
            df = pd.read_excel(os.path.join(path_check_file, fn_check_file), sheet_name=sheet_name, )
        else:
            df = pd.read_excel(os.path.join(path_check_file, fn_check_file), sheet_name=sheet_name, )
        if debug: 
            print("read_check_file:", df.shape, df.columns)
            print(col_name, "col_name in df.columns:", col_name in df.columns)
        if df.shape[1] == 1:
            #logger.info(f"Check file read: shape: " + str(df.shape) )
            logger.info(f"Проверяемый файл '{fn_check_file}': размерность: " + str(df.shape) )
            if  col_name not in df.columns:
                old_col_name = df.columns[0]
                df.rename(columns = {old_col_name: col_name}, inplace=True)
                logger.info(f"Проверяемый файл '{fn_check_file}': Колонка '{old_col_name}' переименована в  '{col_name}'")
        elif col_name in df.columns:
            df = df[[col_name]]
            #logger.info(f"Check file read: shape: " + str(df.shape) )
            logger.info(f"Проверяемый файл '{fn_check_file}': размерность: " + str(df.shape) )
            
        else:
            # logger.error(f"Check file: Not found need column: '{col_name}'")
            logger.error(f"Проверяемый файл '{fn_check_file}', лист '{sheet_name}': не найдена колонка: '{col_name}'")
            read_ok = False
            df = None
    except Exception as err:
        logger.error(f"Проверяемый файл '{fn_check_file}': {err}")
        read_ok = False
        df = None
        # if f"Worksheet named '{sheet_name}' not found" in err:

    return df, read_ok

def read_test_dictionary(path_dict_file, fn_dict_file, sheet_name, col_names):
    df = None
    read_ok = True
    try:
        if sheet_name is None:
            wb = load_workbook(os.path.join(path_dict_file, fn_dict_file))
            for sh_n in wb.get_sheet_names():
                if 'справ' in sh_n.lower():
                    sheet_name = sh_n
            if sheet_name is None: sheet_name = wb.get_sheet_names()[0]
            df = pd.read_excel(os.path.join(path_dict_file, fn_dict_file), sheet_name=sheet_name, )
        else:
            df = pd.read_excel(os.path.join(path_dict_file, fn_dict_file), sheet_name=sheet_name, )
        if df.shape[1] == 2:
            if set(col_names).issubset(df.columns):
                # logger.info(f"Dictionary file read: shape: " + str(df.shape) )
                logger.info(f"Файл локального справочника '{fn_dict_file}' загружен: размерность: " + str(df.shape) )
                # df.rename(columns = {df.columns[0]: col_name}, inplace=True)
            else:
                # logger.error(f"Dictionary file: Not found need columns: '{col_names}'")
                logger.error(f"Файл локального справочника '{fn_dict_file}': не найдены нужные колонки: '{col_names}'")
                read_ok = False
                df = None
        elif set(col_names).issubset(df.columns):
            df = df[col_names]
            logger.info(f"Файл локального справочника '{fn_dict_file}' загружен: размерность: " + str(df.shape) )
        else:
            logger.error(f"Файл локального справочника '{fn_dict_file}': не найдены нужные колонки: '{col_names}'")
            read_ok = False
            df = None
    except Exception as err:
        logger.error(f"Файл локального справочника '{fn_dict_file}': {err}")
        read_ok = False
        df = None
    return df, read_ok

def fuzzy_search (df_test, col_name_check, df_dict, name_col_dict_local, code_col_dict_local, new_cols_fuzzy, similarity_threshold, max_sim_entries=2, n_rows=np.inf):
    def get_code_name(tuple_name_sim_lst, name_col_dict, code_col_dict, similarity_threshold):
        # name = dict_lst[id]
        # values = df_dict.query("@name_col_dict=@name")[[name_col_dict, code_col_dict]].values
        rez = [] # np.array([[], [], []])
        for name, similarity in tuple_name_sim_lst:
            # print("name, similarity", name, similarity, similarity_threshold*100)
            if similarity >= similarity_threshold*100:
                values = df_dict[df_dict[name_col_dict_local]==name][[code_col_dict_local, name_col_dict_local]].values
                # print("values", values)
                rez.append(np.array([similarity, np_unique_nan(values[:,0]), np_unique_nan(values[:,1])], dtype=object))
            else: break
        return np.array(rez)

    
    dict_lst = df_dict[name_col_dict_local].unique()
    # dict_lst[:2]
    df_test[new_cols_fuzzy] = None
    for i_row, row in tqdm(df_test.iterrows(), total = df_test.shape[0]):
        if i_row > n_rows: break
        s = row[col_name_check]
        tuple_name_sim_lst = process.extract(s, dict_lst, limit=max_sim_entries)
        # print(i_row, s, tuple_name_sim_lst)
        values = get_code_name(tuple_name_sim_lst, code_col_dict_local, name_col_dict_local, similarity_threshold)
        # print()
        # print(values)
        if len(values)>0:
            df_test.loc[i_row, new_cols_fuzzy] = np_unique_nan(values[:,0]), np_unique_nan(values[:,1]), np_unique_nan(values[:,2]) 
    return df_test


def fuzzy_search_02 (df_test, col_name_check, df_dict, name_col_dict_local, code_col_dict_local, new_cols_fuzzy, similarity_threshold, max_sim_entries=2, n_rows=np.inf):
    def get_code_name(tuple_name_sim_lst, name_col_dict, code_col_dict, similarity_threshold):
        # name = dict_lst[id]
        # values = df_dict.query("@name_col_dict=@name")[[name_col_dict, code_col_dict]].values
        rez = [] # np.array([[], [], []])
        for name, similarity in tuple_name_sim_lst:
            # print("name, similarity", name, similarity, similarity_threshold*100)
            if similarity >= similarity_threshold*100:
                values = df_dict[df_dict[name_col_dict_local]==name][[code_col_dict_local, name_col_dict_local]].values
                # print("values", values)
                rez.append(np.array([similarity, np_unique_nan(values[:,0]), np_unique_nan(values[:,1])], dtype=object))
            else: break
        return np.array(rez)
    def get_code_name_02(tuple_name_sim_lst, name_col_dict, code_col_dict, similarity_threshold):
        rez = []
        for name, similarity in tuple_name_sim_lst:
            if similarity >= similarity_threshold*100:
                values = df_dict[df_dict[name_col_dict_local]==name][[code_col_dict_local, name_col_dict_local]].values
                rez.append(np.array([similarity, values[:,0], values[:,1]], dtype=object))
            else: break
        return np.array(rez)
    
    dict_lst = df_dict[name_col_dict_local].unique()
    df_test[new_cols_fuzzy] = None
    empty_values_row = np.array([None, None, None])
    dict_test_f02 = {}
    for i_row, row in tqdm(df_test.iterrows(), total = df_test.shape[0]):
        if i_row > n_rows: break
        s = row[col_name_check]
        tuple_name_sim_lst = process.extract(s, dict_lst, limit=max_sim_entries)
        # print(i_row, s, tuple_name_sim_lst)
        values = get_code_name_02(tuple_name_sim_lst, code_col_dict_local, name_col_dict_local, similarity_threshold)
        d_row = { f"{col_name_check}": s}
        for col in new_cols_fuzzy:
            d_row.update({f"{col}":[]})
        if len(values)>0:
            df_test.loc[i_row, new_cols_fuzzy] = np_unique_nan(values[:,0]), np_unique_nan(values[:,1]), np_unique_nan(values[:,2]) 
            
            for i_vr, values_row in enumerate(values):
                # df_test_f02.loc[len(df_test_f02)] = np.hstack((np.array(s, dtype=str), values_row[0], values_row[1] if type(values_row[1])==str else None, values_row[2]))
                for ic, col in enumerate(new_cols_fuzzy):
                    if ic == 0:
                        d_row[f"{col}"].append(values_row[ic])
                    else: d_row[f"{col}"].append(values_row[ic][0])
            dict_test_f02[i_row] = d_row
        elif len(values)==0:
            # df_test_f02.loc[len(df_test_f02)] = np.hstack((np.array(s, dtype=str), empty_values_row))
            dict_test_f02[i_row] = d_row
        
    return df_test, dict_test_f02
new_cols_fuzzy = ['similarity_fuzzy', 'sim_fuzzy_code', 'sim_fuzzy_name', ]
new_cols_semantic = ['sim_semantic_1_local', 'code_semantic_1_local', 'name_semantic_1_local']
new_cols_semantic_gos = ['sim_semantic_2_gos', 'code_semantic_2_gos', 'name_semantic_2_gos']
new_cols_semantic_national = ['sim_semantic_3_national', 'code_semantic_3_national', 'name_semantic_3_national']
new_cols_semantic_gos_options = ['sim_semantic_4_gos_option', 'option_semantic_4_gos_option', 'code_semantic_4_gos_option', 'name_semantic_4_gos_option']
# similarity_threshold = .8
# max_sim_entries = 2

def save_stat(df_test, data_processed_dir, fn_check_file, max_sim_entries, similarity_threshold):
    nums_lst = []
    total_num_recs = df_test.shape[0]
    # nums_lst.append(['total_num_recs', total_num_recs])
    nums_lst.append(['Общее количество записей в наборе', total_num_recs])
    # mask_cols = [df_test.columns[0]]
    mask_cols = []
    try:
        num_found_rec_fuzzy = df_test[df_test['sim_fuzzy_name'].notnull()].shape[0]
        # nums_lst.append(['num_found_rec_fuzzy', num_found_rec_fuzzy])
        nums_lst.append(['Найдено записей по fuzzy поиску', num_found_rec_fuzzy])
        mask_cols.append('sim_fuzzy_name')
    except:
        num_found_rec_fuzzy = None
    try:
        num_found_rec_semantic_1_local = df_test[df_test['name_semantic_1_local'].notnull()].shape[0]
        # nums_lst.append(['num_found_rec_semantic_1_local', num_found_rec_semantic_1_local])
        nums_lst.append(['Найдено записей по семантическому поиску по локальному справочнику', num_found_rec_semantic_1_local])
        mask_cols.append('name_semantic_1_local')
    except:
        num_found_rec_semantic_1_local = None
    # print(f"num_found_rec_semantic_1_local: {num_found_rec_semantic_1_local}")
    try:
        num_found_rec_semantic_2_gos = df_test[df_test['name_semantic_2_gos'].notnull()].shape[0]
        # nums_lst.append(['num_found_rec_semantic_2_gos', num_found_rec_semantic_2_gos])
        nums_lst.append(['Найдено записей по семантическому поиску по гос реестру МИ', num_found_rec_semantic_2_gos])
        mask_cols.append('name_semantic_2_gos')
    except:
        num_found_rec_semantic_2_gos = None
    try:
        num_found_rec_semantic_3_national = df_test[df_test['name_semantic_3_national'].notnull()].shape[0]
        # nums_lst.append(['num_found_rec_semantic_3_national', num_found_rec_semantic_3_national])
        nums_lst.append(['Найдено записей по семантическому поиску по национальному реестру МИ', num_found_rec_semantic_3_national])
        mask_cols.append('name_semantic_3_national')
    except:
        num_found_rec_semantic_3_national = None
    try:
        num_found_rec_semantic_4_gos_option = df_test[df_test['option_semantic_4_gos_option'].notnull()].shape[0]
        # nums_lst.append(['num_found_rec_semantic_4_gos_option', num_found_rec_semantic_4_gos_option])
        nums_lst.append(['Найдено записей по семантическому поиску по Вариантам исполнения гос реестра МИ', num_found_rec_semantic_2_gos])
        mask_cols.append('option_semantic_4_gos_option')
    except:
        num_found_rec_semantic_4_gos_option = None
    for ic, col in enumerate(mask_cols):
        if ic == 0:
            mask_not_found = df_test[col].isnull()
        else: mask_not_found = mask_not_found & df_test[col].isnull()
    num_found_rec_total = df_test[~mask_not_found].shape[0]
    # nums_lst.append(['num_found_rec_total', num_found_rec_total])
    nums_lst.append(['Всего найдено записей', num_found_rec_total])

    print(f"Общее количество записей в наборе: {total_num_recs}",
          f"\nНайдено записей по fuzzy поиску: {num_found_rec_fuzzy}" if num_found_rec_fuzzy is not None else '',
          f"\nНайдено записей по семантическому поиску по локальному справочнику: {num_found_rec_semantic_1_local}" if num_found_rec_semantic_1_local is not None else '',
          f"\nНайдено записей по семантическому поиску по гос реестру МИ: {num_found_rec_semantic_2_gos}" if num_found_rec_semantic_2_gos is not None else '', 
          f"\nНайдено записей по семантическому поиску по национальному реестру МИ: {num_found_rec_semantic_3_national}" if num_found_rec_semantic_3_national is not None else '',
          f"\nНайдено записей по семантическому поиску по Вариантам исполнения гос реестра МИ: {num_found_rec_semantic_4_gos_option}" if num_found_rec_semantic_4_gos_option is not None else '',
          f"\nВсего найдено записей: {num_found_rec_total}",
          '')
    
    wb = Workbook()
    ws = wb.active
    # ws = wb['Statistics']
    ws.title = 'Statistics'
    alignment=Alignment(horizontal='left', #'general',
                             vertical= 'top', #'bottom',
                             text_rotation=0,
                             wrap_text=True,
                             shrink_to_fit=False,
                             indent=0)
    # cell = ws.cell(row=ir+1, column=ic+1 + desc_cols_num)
    # cell.comment = None
    cell = ws['A1']
    cell.alignment = alignment
    ws.column_dimensions[cell.column_letter].width = 40
    cell = ws['B1']
    alignment=Alignment(horizontal='right', #'general',
                             vertical= 'top', #'bottom',
                             text_rotation=0,
                             wrap_text=True,
                             shrink_to_fit=False,
                             indent=0)
    cell.alignment = alignment
    ws.column_dimensions[cell.column_letter].width = 10

    # ws['A2'], ws['B2'] = 'similarity_threshold, %', similarity_threshold
    ws['A1'], ws['B1'] = 'Минимальный % сходства', similarity_threshold
    ws["B1"].number_format = "0%"
    # ws['A1'], ws['B1'] = 'max_sim_entries', max_sim_entries
    ws['A2'], ws['B2'] = 'Максимальное количество уникальных наименований', max_sim_entries
    
    for lst in nums_lst:
        ws.append(lst)
    fn_main = fn_check_file.split('.xlsx')[0] + '_stat'
    offset = datetime.timezone(datetime.timedelta(hours=3))
    dt = datetime.datetime.now(offset)
    str_date = dt.strftime("%Y_%m_%d_%H%M")
    fn = fn_main + '_' + str_date + '.xlsx'
    wb.save(os.path.join(data_processed_dir,fn))
    logger.info(f"Файл статистики: '{fn}' созранен в '{data_processed_dir}'")
    return fn


def lst_2_s(lst):
    if not (type(lst)==str) and ((type(lst)==list) or (type(lst)==np.ndarray)):
        if len(lst)==1: rez = lst[0]
        else: rez = lst
    else: rez = lst[0]
    return rez

def semantic_search (df_test, col_name_check, 
                     dict_unique, df_dict, name_col_dict, code_col_dict,
                     option_col_dict,
                     model, dict_embedding, 
                     new_cols_semantic, 
                     similarity_threshold, max_sim_entries=2, n_rows=np.inf,
                     debug=False):
    def get_code_name(dict_id_score_lst, dict_unique, df_dict, name_col_dict, code_col_dict, option_col_dict, similarity_threshold, debug=False):
        # name = dict_lst[id]
        # values = df_dict.query("@name_col_dict=@name")[[name_col_dict, code_col_dict]].values
        rez = [] # np.array([[], [], []])
        # 'corpus_id': 182, 'score'
        if len(dict_id_score_lst[0]) > 0:
            for dict_id_score in dict_id_score_lst[0]:
                if debug: print("get_code_name: dict_score_id", dict_id_score)
                id, score = dict_id_score.values()
                if debug: print(f"get_code_name: score: {score}, id : {id}")
                if float(score) >= similarity_threshold:
                    name = dict_unique[id]
                    if option_col_dict is None:
                        values = df_dict[df_dict[name_col_dict]==name][[code_col_dict, name_col_dict]].values
                        try:
                            rez.append(np.array([round(score*100), np_unique_nan(values[:,0]), np_unique_nan(values[:,1])], dtype=object))
                        except Exception as err:
                            if debug: print("get_code_name: ", err, values.shape, values)
                            rez.append(np.array([round(score*100), values[:,0], values[:,1]], dtype=object))
                    else:
                        values = df_dict[df_dict[option_col_dict]==name][[option_col_dict, code_col_dict, name_col_dict]].values
                        try:
                            rez.append(np.array([round(score*100), np_unique_nan(values[:,0]), 
                                             np_unique_nan(values[:,1]), np_unique_nan(values[:,2])], dtype=object))
                        except Exception as err:
                            if debug: print("get_code_name: ", err, values.shape, values)
                            rez.append(np.array([round(score*100), values[:,0], values[:,1], values[:,2]], dtype=object))
                else: break
            
            return np.array(rez)
        else: 
            if option_col_dict is None: return np.array([None, None, None])
            else: return np.array([None, None, None, None])

    df_test[new_cols_semantic] = None
    for i_row, row in tqdm(df_test.iterrows(), total = df_test.shape[0]):
    # for i_row, row in df_test.iterrows():
        if i_row > n_rows: break
        s = row[col_name_check]
        # tuple_name_sim_lst = process.extract(s, dict_lst, limit=max_sim_entries)
        query_embedding = model.encode(s) #row[col_name_check])
        if debug: print("\nsemantic_search:", i_row, s)
        dict_id_score_lst = util.semantic_search (query_embedding, dict_embedding, top_k = max_sim_entries)
        if debug: print(f"semantic_search: dict_id_score_lst: {dict_id_score_lst}")
        
        values = get_code_name(dict_id_score_lst, dict_unique, df_dict, name_col_dict, code_col_dict, option_col_dict, similarity_threshold, debug=debug)
        if debug: 
            print("semantic_search: -> get_code_name values:", values )
            print()
        # df_test.loc[i_row, ['sim_3_semantic_multy_lang_test', 'name_3_semantic_multy_lang_test', 'sim_4_semantic_multy_lang_test', 'name_4_semantic_multy_lang_test']] = \
        #         round(rez[0][0]['score']*100), dict_lst[rez[0][0]['corpus_id']], round(rez[0][1]['score']*100), dict_lst[rez[0][1]['corpus_id']]
        # Параметр limit по умолчанию имеет значение 5
        
        if len(values)>0:
            if option_col_dict is None:
                try:
                    df_test.loc[i_row, new_cols_semantic] = lst_2_s(values[:,0]), np_unique_nan(values[:,1]), np_unique_nan(values[:,2])
                except Exception as err:
                    if debug: print("semantic_search:", err, values.shape, values)
                    df_test.loc[i_row, new_cols_semantic] = lst_2_s(values[:,0]), values[:,1], values[:,2]
            else: 
                # print(values.shape, values)
                try:
                    df_test.loc[i_row, new_cols_semantic] = lst_2_s(values[:,0]), np_unique_nan(values[:,1]), np_unique_nan(values[:,2]), np_unique_nan(values[:,3])
                except Exception as err:
                    if debug: print("semantic_search:", err, values.shape, values)
                    df_test.loc[i_row, new_cols_semantic] = lst_2_s(values[:,0]), values[:,1], values[:,2], values[:,3]

    return df_test

def semantic_search_02 (df_test, col_name_check, 
                     dict_unique, df_dict, name_col_dict, code_col_dict,
                     option_col_dict,
                     model, dict_embedding, 
                     new_cols_semantic, 
                     similarity_threshold, max_sim_entries=2, n_rows=np.inf,
                     debug=False):
    def get_code_name(dict_id_score_lst, dict_unique, df_dict, name_col_dict, code_col_dict, option_col_dict, similarity_threshold, debug=False):
        # name = dict_lst[id]
        # values = df_dict.query("@name_col_dict=@name")[[name_col_dict, code_col_dict]].values
        rez = [] # np.array([[], [], []])
        # 'corpus_id': 182, 'score'
        if len(dict_id_score_lst[0]) > 0:
            for dict_id_score in dict_id_score_lst[0]:
                if debug: print("get_code_name: dict_score_id", dict_id_score)
                id, score = dict_id_score.values()
                if debug: print(f"get_code_name: score: {score}, id : {id}")
                if float(score) >= similarity_threshold:
                    name = dict_unique[id]
                    if option_col_dict is None:
                        values = df_dict[df_dict[name_col_dict]==name][[code_col_dict, name_col_dict]].values
                        try:
                            rez.append(np.array([round(score*100), np_unique_nan(values[:,0]), np_unique_nan(values[:,1])], dtype=object))
                        except Exception as err:
                            if debug: print("get_code_name: ", err, values.shape, values)
                            rez.append(np.array([round(score*100), values[:,0], values[:,1]], dtype=object))
                    else:
                        values = df_dict[df_dict[option_col_dict]==name][[option_col_dict, code_col_dict, name_col_dict]].values
                        try:
                            rez.append(np.array([round(score*100), np_unique_nan(values[:,0]), 
                                             np_unique_nan(values[:,1]), np_unique_nan(values[:,2])], dtype=object))
                        except Exception as err:
                            if debug: print("get_code_name: ", err, values.shape, values)
                            rez.append(np.array([round(score*100), values[:,0], values[:,1], values[:,2]], dtype=object))
                else: break
            
            return np.array(rez)
        else: 
            if option_col_dict is None: return np.array([None, None, None])
            else: return np.array([None, None, None, None])
    def get_code_name_02(dict_id_score_lst, dict_unique, df_dict, name_col_dict, code_col_dict, option_col_dict, similarity_threshold, debug=False):
        rez = [] # np.array([[], [], []])
        if len(dict_id_score_lst[0]) > 0:
            for dict_id_score in dict_id_score_lst[0]:
                if debug: print("get_code_name: dict_score_id", dict_id_score)
                id, score = dict_id_score.values()
                if debug: print(f"get_code_name: score: {score}, id : {id}")
                if float(score) >= similarity_threshold:
                    name = dict_unique[id]
                    if option_col_dict is None:
                        # values = df_dict[df_dict[name_col_dict]==name][[code_col_dict, name_col_dict]].values
                        values = df_dict[df_dict[name_col_dict]==name][[code_col_dict, name_col_dict]].\
                                  drop_duplicates(subset=[code_col_dict, name_col_dict]).values
                        try:
                            # rez.append(np.array([round(score*100), np_unique_nan(values[:,0]), np_unique_nan(values[:,1])], dtype=object))
                            rez.append(np.array([round(score*100), lst_2_s(values[:,0]), lst_2_s(values[:,1])], dtype=object))
                        except Exception as err:
                            rez.append(np.array([round(score*100), lst_2_s(values[:,0]), lst_2_s(values[:,1])], dtype=object))
                            if debug: 
                                print("get_code_name_02: ", err, values.shape, values)
                                print("get_code_name_02: rez:", rez)
                    else:
                        # values = df_dict[df_dict[option_col_dict]==name][[option_col_dict, code_col_dict, name_col_dict]].values
                        values = df_dict[df_dict[option_col_dict]==name][[option_col_dict, code_col_dict, name_col_dict]].\
                              drop_duplicates(subset=[option_col_dict, code_col_dict, name_col_dict]).values
                        try:
                            # rez.append(np.array([round(score*100), values[:,0], values[:,1], values[:,2]], dtype=object))
                            rez.append(np.array([round(score*100), lst_2_s(values[:,0]), lst_2_s(values[:,1]), lst_2_s(values[:,2])], dtype=object))
                        except Exception as err:
                            # print("get_code_name_02: ", err, values.shape, values)
                            rez.append(np.array([round(score*100), lst_2_s(values[:,0]), lst_2_s(values[:,1]), lst_2_s(values[:,2])], dtype=object))
                            if debug: 
                                print("get_code_name_02: ", err, values.shape, values)
                                print("get_code_name_02: rez:", rez)
                            
                else: break
            
            return np.array(rez)
        else: 
            if option_col_dict is None: return np.array([None, None, None])
            else: return np.array([None, None, None, None])

    df_test[new_cols_semantic] = None
    if option_col_dict is None: empty_values_row = np.array([None, None, None])
    else: empty_values_row = np.array([None, None, None, None])
    # df_test_f02 = pd.DataFrame( (1 + len(new_cols_semantic))*[], columns = [col_name_check] + new_cols_semantic)
    # print(df_test_f02.columns)
    dict_test_f02 = {}

    for i_row, row in tqdm(df_test.iterrows(), total = df_test.shape[0]):
    
        if i_row > n_rows: break
        s = row[col_name_check]
        d_row = { f"{col_name_check}": s}
        for col in new_cols_semantic:
            d_row.update({f"{col}":[]})
        
        query_embedding = model.encode(s) #row[col_name_check])
        # if debug: print("\nsemantic_search:", i_row, s)
        dict_id_score_lst = util.semantic_search (query_embedding, dict_embedding, top_k = max_sim_entries)
        # if debug: print(f"semantic_search: dict_id_score_lst: {dict_id_score_lst}")
        
        # values = get_code_name(dict_id_score_lst, dict_unique, df_dict, name_col_dict, code_col_dict, option_col_dict, similarity_threshold, debug=debug)
        values = get_code_name_02(dict_id_score_lst, dict_unique, df_dict, name_col_dict, code_col_dict, option_col_dict, similarity_threshold, debug=debug)
        if debug: 
            #print("semantic_search: -> get_code_name values:", values )
            #print()
            pass
        if len(values)>0:
            if option_col_dict is None:
                try:
                    df_test.loc[i_row, new_cols_semantic] = lst_2_s(values[:,0]), np_unique_nan(values[:,1]), np_unique_nan(values[:,2])
                    for i_vr, values_row in enumerate(values):
                        # if i_row==0: print(values_row.shape, values_row)
                        # df_test_f02.loc[len(df_test_f02)] = np.hstack((np.array(s, dtype=str), values_row[0], values_row[1] if type(values_row[1])==str else None, values_row[2]))
                        for ic, col in enumerate(new_cols_semantic):
                            if ic == 0:
                                if ((type(values_row[ic])==list) or (type(values_row[ic])==np.ndarray)):
                                    # if ((type(d_row[f"{col}"])==list)):
                                    d_row[f"{col}"].extend(lst_2s(values_row[ic]))
                                    # else:
                                    #     d_row[f"{col}"].append(lst_2s(values_row[ic]))
                                else:
                                    d_row[f"{col}"].append(values_row[ic])
                            else: 
                                # d_row[f"{col}"].append(values_row[ic][0])
                                if ((type(values_row[ic])==list) or (type(values_row[ic])==np.ndarray)):
                                    d_row[f"{col}"].extend(lst_2s(values_row[ic]))
                                else:
                                    d_row[f"{col}"].append(values_row[ic])
                            # else: d_row[f"{col}"].append(values_row[ic])
                    dict_test_f02[i_row] = d_row
                    
                    # lst_test_f02[i_row].update(dict(zip(new_cols_semantic,(values_row[0], values_row[1][0] if type(values_row[1][0])==str else None ,values_row[2][0]))))
                except Exception as err:
                    if debug: 
                        print()
                        print("\nsemantic_search:", i_row, s)
                        print("semantic_search:", err, values.shape, values)
                    # print("semantic_search:", err, values.shape, values)
                    df_test.loc[i_row, new_cols_semantic] = lst_2_s(values[:,0]), values[:,1], values[:,2]
                    for i_vr, values_row in enumerate(values):
                        for ic, col in enumerate(new_cols_semantic):
                            if ic == 0:
                                if ((type(values_row[ic])==list) or (type(values_row[ic])==np.ndarray)):
                                    # if ((type(d_row[f"{col}"])==list)):
                                    d_row[f"{col}"].extend(values_row[ic])
                                    # else:
                                    #     d_row[f"{col}"].append(values_row[ic])
                                else:
                                    d_row[f"{col}"].append(values_row[ic])
                            else: 
                                if ((type(values_row[ic])==list) or (type(values_row[ic])==np.ndarray)):
                                    d_row[f"{col}"].extend(values_row[ic])
                                else:
                                    d_row[f"{col}"].append(values_row[ic])
                                # d_row[f"{col}"].append(values_row[ic][0])
                    dict_test_f02[i_row] = d_row
            else: 
                # print(values.shape, values)
                try:
                    df_test.loc[i_row, new_cols_semantic] = lst_2_s(values[:,0]), np_unique_nan(values[:,1]), np_unique_nan(values[:,2]), np_unique_nan(values[:,3])
                    # lst_test_f02[i_row].update(dict(zip(new_cols_semantic,(values_row[0], values_row[1][0] if type(values_row[1][0])==str else None , values_row[2][0], values_row[3][0]))))
                    for i_vr, values_row in enumerate(values):
                        # if i_row==0: print(values_row.shape, values_row)
                        # df_test_f02.loc[len(df_test_f02)] = np.hstack((np.array(s, dtype=str), values_row[0], values_row[1] if type(values_row[1])==str else None, values_row[2], values_row[3]))
                        for ic, col in enumerate(new_cols_semantic):
                            # if ic == 0:
                            #     d_row[f"{col}"].append(values_row[ic])
                            # else: d_row[f"{col}"].append(values_row[ic][0])
                            # else: d_row[f"{col}"].append(values_row[ic])
                            if ic == 0:
                                if ((type(values_row[ic])==list) or (type(values_row[ic])==np.ndarray)):
                                    d_row[f"{col}"].extend(values_row[ic])
                                else:
                                    d_row[f"{col}"].append(values_row[ic])
                            else: 
                                if ((type(values_row[ic])==list) or (type(values_row[ic])==np.ndarray)):
                                    d_row[f"{col}"].extend(values_row[ic])
                                else:
                                    d_row[f"{col}"].append(values_row[ic])
                    dict_test_f02[i_row] = d_row
                except Exception as err:
                    if debug: 
                        print()
                        print("\nsemantic_search:", i_row, s)
                        print("semantic_search:", err, values.shape, values)
                    df_test.loc[i_row, new_cols_semantic] = lst_2_s(values[:,0]), values[:,1], values[:,2], values[:,3]
                    for i_vr, values_row in enumerate(values):
                        for ic, col in enumerate(new_cols_semantic):
                            # if ic == 0:
                            #     d_row[f"{col}"].append(values_row[ic])
                            # else: d_row[f"{col}"].append(values_row[ic][0])
                            if ic == 0:
                                if ((type(values_row[ic])==list) or (type(values_row[ic])==np.ndarray)):
                                    d_row[f"{col}"].extend(values_row[ic])
                                else:
                                    d_row[f"{col}"].append(values_row[ic])
                            else: 
                                if ((type(values_row[ic])==list) or (type(values_row[ic])==np.ndarray)):
                                    d_row[f"{col}"].extend(values_row[ic])
                                else:
                                    d_row[f"{col}"].append(values_row[ic])
                            # else: d_row[f"{col}"].append(values_row[ic])
                    dict_test_f02[i_row] = d_row
        elif len(values)==0:
            # df_test_f02.loc[len(df_test_f02)] = np.hstack((np.array(s, dtype=str), empty_values_row))
            # lst_test_f02[i_row].update(dict(zip(new_cols_semantic, empty_values_row)))
            dict_test_f02[i_row] = d_row
        else: dict_test_f02[i_row] = d_row
            # pass
    return df_test, dict_test_f02

def load_sentence_model():
    # logger.info(f"MultyLangual model dounlowd - start...")
    logger.info(f"Загрузка предобученной нейросети - Старт...")
    model = SentenceTransformer('multi-qa-MiniLM-L6-cos-v1')
    # logger.info(f"MultyLangual model dounlowd - done!")
    logger.info(f"Загрузка предобученной нейросети - Финиш!")
    return model

fn_list = []
    
def def_form(fn_list):
    fn_check_file_drop_douwn = widgets.Dropdown( options=fn_list, value=fn_list[0] if len(fn_list) > 0 else None, disabled=False)
    fn_dict_file_drop_douwn = widgets.Dropdown( options= [None] + fn_list, value= None, disabled=False, )
    radio_btn_big_dict = widgets.RadioButtons(options=['Да', 'Нет'], value= 'Да', disabled=False) # description='Check me',    , indent=False
    radio_btn_prod_options = widgets.RadioButtons(options=['Да', 'Нет'], value= 'Нет', disabled=False if radio_btn_big_dict.value=='Да' else True )
    similarity_threshold_slider = widgets.IntSlider(min=1,max=100, value=90)
    max_entries_slider = widgets.IntSlider(min=1,max=5, value=2)

    form_item_layout = Layout(display='flex', flex_flow='row', justify_content='space-between')
    check_box = Box([Label(value='Проверяемый файл:'), fn_check_file_drop_douwn], layout=form_item_layout) 
    dict_box = Box([Label(value='Файл справочника:'), fn_dict_file_drop_douwn], layout=form_item_layout) 
    big_dict_box = Box([Label(value='Использовать большие справочники:'), radio_btn_big_dict], layout=form_item_layout) 
    prod_options_box = Box([Label(value='Искать в Вариантах исполнения (+10 мин):'), radio_btn_prod_options], layout=form_item_layout) 
    similarity_threshold_box = Box([Label(value='Минимальный % сходства позиций:'), similarity_threshold_slider], layout=form_item_layout) 
    max_entries_box = Box([Label(value='Максимальное кол-во найденных уникальных позиций , %:'), max_entries_slider], layout=form_item_layout) 
    
    form_items = [check_box, dict_box, big_dict_box, prod_options_box, similarity_threshold_box, max_entries_box]
    
    form = Box(form_items, layout=Layout(display='flex', flex_flow= 'column', border='solid 2px', align_items='stretch', width='50%')) #width='auto'))
    return form, fn_check_file_drop_douwn, fn_dict_file_drop_douwn, radio_btn_big_dict, radio_btn_prod_options, similarity_threshold_slider, max_entries_slider
    # form = Box(form_items, layout=Layout(display='flex', flex_flow= 'column', border='solid 2px', align_items='stretch', width='70%')) #width='auto'))
def on_big_dict_value_change(change):
    global radio_btn_prod_options
    if change.new == 'Да':
        radio_btn_prod_options.disabled = False 
        # radio_btn_prod_options.value = 'Нет' 
    else:
        radio_btn_prod_options.disabled = True 
        radio_btn_prod_options.value = 'Нет'     

def dict_2_df(dict_test_f02, col_name_check, new_cols_fuzzy, new_cols_semantic,
             new_cols_semantic_gos, new_cols_semantic_national, new_cols_semantic_gos_options):
    all_columns = [col_name_check] + new_cols_fuzzy + new_cols_semantic\
             + new_cols_semantic_gos + new_cols_semantic_national + new_cols_semantic_gos_options
    
    # df_total = pd.DataFrame( [len(all_columns)*[None]],  columns=all_columns)
    df_total = pd.DataFrame( None,  columns=all_columns)
    # display(df_total)
    # print(100*'*')
    for i_row, (k,v) in enumerate(dict_test_f02.items()):
        # dict_test_f02[k].update(dict_test_f02_tmp[k])
        # for kk,vv in list(dict_test_f02[k].items())[1:]:
        for kk,vv in list(dict_test_f02[k].items())[:]:
            if not ((type(vv)==list) or (type(vv)==np.ndarray)):
                dict_test_f02[k][kk] = [vv]
            # if ((type(vv)==list) or (type(vv)==np.ndarray)):
            #     dict_test_f02[k][kk] = vv
            # else:
            #     dict_test_f02[k][kk] = [vv]
        try:
            d = dict_test_f02[k]
            # print("len(d):", len(d))
            # l = list(zip_longest(*list(d.values())[1:]))
            l = list(zip_longest(*list(d.values())[:]))
            # l = [[d['Наименование МИ/РМ']] + list(ll) for ll in l]
            # l = [d['Наименование МИ/РМ'] + list(ll) for ll in l]
            l = [list(ll) for ll in l]
            df = pd.DataFrame(l, columns = d.keys())
            # df = pd.DataFrame(l, columns = ['Наименование МИ/РМ'] + list(d.keys()))
            # df = pd.DataFrame.from_dict(dict_test_f02[k], orient='columns') #, index=[0]
        except Exception as err:
            print(100*'*')
            print(err)
            print(i_row, dict_test_f02[k])
            # display(df)
            # sys.exit(2)
        # print("df.columns:", df.columns)
        df_total = pd.concat([df_total, df])
        
    # print(df_total.shape, df.shape)
    # print(len(dict_test_f02))
    df_total = df_total.iloc[1:]
    return df_total


df_mi_org_gos, df_mi_org_gos_prod_options, df_mi_national, dict_embedding_gos_multy, dict_embedding_gos_prod_options_multy, dict_embedding_national_multy, dict_lst_gos_prod_options = None, None, None, None, None, None, None

def mi_search( data_source_dir, data_processed_dir,
              fn_check_file, sheet_name_check, col_name_check,
              fn_dict_file, sheet_name_dict, name_col_dict_local, code_col_dict_local,
              similarity_threshold = 0.9,
              max_sim_entries=2,
              by_big_dict = False,
              by_prod_options = False,
              df_dicts = [],
              debug=False,
              n_rows = np.inf
):
    # if 'data_source_dir' not in locals()  or 'data_processed_dir' not in locals():
    #     logger.error(F"Переменные data_source_dir' и'data_processed_dir' не определены")
    #     logger.info(F"Перезапустите п.3. Create Data Direcoties")
    #     sys.exit(2)
    
    if fn_dict_file is None and not by_big_dict:
        by_big_dict = True
    test_ok = test_inputs(data_source_dir, 
              fn_check_file, sheet_name_check, col_name_check,
              fn_dict_file, sheet_name_dict, [name_col_dict_local, code_col_dict_local], by_big_dict,
              )
    if not test_ok:
        # logger.error(f"Error of input params - Program is over")
        logger.error(f"Ошибка входных параметров - Работа программы завершена")
        sys.exit(2)
    
    df_test, read_ok_check = read_check_file(data_source_dir, fn_check_file, sheet_name_check, col_name_check)
    if read_ok_check: display(df_test.head(2))
    if fn_dict_file is not None:
        df_dict, read_ok_dict = df_dict, read_ok_dict = read_test_dictionary(data_source_dir, fn_dict_file, sheet_name_dict, [name_col_dict_local, code_col_dict_local]) #col_names_dict) 
        if read_ok_dict: display(df_dict.head(2))
        else: print(type(df_dict))
    else: read_ok_dict = True
    if not read_ok_check or (not read_ok_dict and not semantic_search):
        # logger.error(f"Error of input data - Program is over")
        logger.error(f"Ошибка в листах или колонках проверяемого файла - Работа программы завершена")
        sys.exit(2)
    
    format_cols = [60]
    format_cols.extend([10, 15, 60])
    if fn_dict_file is not None and df_dict is not None:
        # logger.info("Fuzzy search on local dictionary - start...")
        logger.info("Нечеткий (Fuzzy) поиск по локальному справочнику - Старт...")
        new_cols_fuzzy = ['similarity_fuzzy', 'sim_fuzzy_code', 'sim_fuzzy_name', ]
        df_test, dict_test_f02 = fuzzy_search_02 (df_test, col_name_check, df_dict, name_col_dict_local, code_col_dict_local, 
                                new_cols_fuzzy, similarity_threshold, max_sim_entries, n_rows = n_rows)
            # fuzzy_search (df_test, col_name_check, df_dict, name_col_dict_local, code_col_dict_local, new_cols_fuzzy, similarity_threshold, max_sim_entries=2, n_rows=np.inf)
        display(df_test.head(2))
        # print("df_test.shape:", df_test.shape)
        # print(list(dict_test_f02.items())[:2])
        # print("len(dict_test_f02):", len(dict_test_f02))
    else: 
        new_cols_fuzzy = []
        dict_test_f02 = None
    
    
    model = load_sentence_model()

    if fn_dict_file is not None and df_dict is not None:
        format_cols.extend([10,15, 60])
        new_cols_semantic = ['sim_semantic_1_local', 'code_semantic_1_local', 'name_semantic_1_local']
        dict_local_unique = df_dict[name_col_dict_local].unique()
        option_col_dict_local = None
        logger.info("Формирование ембеддингов локального справочника - Старт..." )
        dict_embedding_local = model.encode(dict_local_unique, show_progress_bar=True)
        logger.info("Формирование ембеддингов локального справочника - Финиш!" )
        # df_test = semantic_search (df_test, col_name_check, 
        # logger.info("Semantic search on local dictionary - start...")
        logger.info("Сематнтический поиск по локальному справочнику - Старт..." )
        df_test, dict_test_f02_tmp = semantic_search_02 (df_test, col_name_check, 
                     dict_local_unique, df_dict, name_col_dict_local, code_col_dict_local,
                     option_col_dict_local, 
                     model, dict_embedding_local, 
                     new_cols_semantic, 
                     similarity_threshold, max_sim_entries, n_rows = n_rows)
        display(df_test.head(2))
        # print(list(dict_test_f02_tmp.items())[:2])
        # print("df_test.shape:", df_test.shape)
        # print("len(dict_test_f02_tmp):", len(dict_test_f02_tmp))
        if dict_test_f02 is not None:
            for k,v in dict_test_f02.items():
                try: #if dict_test_f02.get(k) is not None:
                    dict_test_f02[k].update(dict_test_f02_tmp.get(k))
                #else:
                except Exception as err:
                    print(err, "dict_test_f02.get(k) is None", k)
        else: dict_test_f02 = dict_test_f02_tmp
        # print("len(dict_test_f02 - after semantic search 01:", len(dict_test_f02))
    else: 
        new_cols_semantic = []
        dict_test_f02 = None
    # print("len(dict_test_f02 - after semantic search 02:", len(dict_test_f02))
    if by_big_dict:
        df_mi_org_gos, df_mi_org_gos_prod_options, df_mi_national, dict_embedding_gos_multy, dict_embedding_gos_prod_options_multy, dict_embedding_national_multy, dict_lst_gos_prod_options = df_dicts
        
        test_ok = test_big_dictionaries(df_mi_org_gos, df_mi_org_gos_prod_options, df_mi_national, dict_embedding_gos_multy, dict_embedding_gos_prod_options_multy, dict_embedding_national_multy, dict_lst_gos_prod_options)
        if not test_ok:
           logger.error(f"Работа программы завершена")
           sys.exit(2)
        else:
            #global df_mi_org_gos, df_mi_org_gos_prod_options, df_mi_national, dict_embedding_gos_multy, dict_embedding_gos_prod_options_multy, dict_embedding_national_multy, dict_lst_gos_prod_options
            #print(df_mi_org_gos.shape, df_mi_org_gos_prod_options.shape, df_mi_national.shape)
            #print(dict_embedding_gos_multy.shape, dict_embedding_gos_prod_options_multy.shape, dict_embedding_national_multy.shape)
            #print(dict_lst_gos_prod_options.shape)
            pass
    
        format_cols.extend([10,15, 60])
        new_cols_semantic_gos = ['sim_semantic_2_gos', 'code_semantic_2_gos', 'name_semantic_2_gos']
        name_col_dict_gos, code_col_dict_gos = 'name_clean', 'kind'
        option_col_dict_gos = None
        # logger.info("Semantic search on gos dictionary - start...")
        logger.info("Сематнтический поиск по гос реестру МИ - Старт..." )
        dict_gos_unique = df_mi_org_gos[name_col_dict_gos].unique()
        # dict_embedding_gos = model.encode(dict_gos_unique, show_progress_bar=True)
        # df_test = semantic_search (df_test, col_name_check, 
        df_test, dict_test_f02_tmp = semantic_search_02 (df_test, col_name_check, 
                     dict_gos_unique, df_mi_org_gos, name_col_dict_gos, code_col_dict_gos,
                     option_col_dict_gos,
                     model, dict_embedding_gos_multy, 
                     new_cols_semantic_gos, 
                     similarity_threshold, max_sim_entries, n_rows = n_rows)
        display(df_test.head(2))
        if dict_test_f02 is not None:
            for k,v in dict_test_f02.items():
                try: #if dict_test_f02.get(k) is not None:
                    dict_test_f02[k].update(dict_test_f02_tmp.get(k))
                #else:
                except Exception as err:
                    print(err, "dict_test_f02.get(k) is None", k)
        else: dict_test_f02 = dict_test_f02_tmp

        format_cols.extend([10,15, 60])
        new_cols_semantic_national = ['sim_semantic_3_national', 'code_semantic_3_national', 'name_semantic_3_national']
        name_col_dict_national, code_col_dict_national = 'name', 'code'
        option_col_dict_national = None
        # logger.info("Semantic search on national dictionary - start...")
        logger.info("Сематнтический поиск по реестру МИ по национальнмоу законодательству - Старт..." )
        dict_national_unique = df_mi_national[name_col_dict_national].unique()
        # dict_embedding_gos = model.encode(dict_gos_unique, show_progress_bar=True)
        # df_test = semantic_search (df_test, col_name_check, 
        df_test, dict_test_f02_tmp = semantic_search_02 (df_test, col_name_check, 
                     dict_national_unique, df_mi_national, name_col_dict_national, code_col_dict_national,
                     option_col_dict_national,
                     model, dict_embedding_national_multy, # dict_embedding_gos_prod_options_multy, #
                     new_cols_semantic_national, 
                     similarity_threshold, max_sim_entries, n_rows = n_rows)
        display(df_test.head(2))
        for k,v in dict_test_f02.items():
            try: # if dict_test_f02.get(k) is not None:
                dict_test_f02[k].update(dict_test_f02_tmp.get(k))
            #else:
            except Exception as err:
                print(err, "dict_test_f02.get(k) is None", k)
        
        if by_prod_options:
            format_cols.extend([10, 60, 15, 60])
            new_cols_semantic_gos_options = ['sim_semantic_4_gos_option', 'option_semantic_4_gos_option', 'code_semantic_4_gos_option', 'name_semantic_4_gos_option']
            name_col_dict_gos_option, code_col_dict_gos_option, option_col_dict_gos_option = 'name_clean', 'kind', 'option'
            # logger.info("Semantic search on options in gos dictionary - start...")
            logger.info("Сематнтический поиск по Вариантам исполнения из гос реестра МИ - Старт..." )
            #dict_gos_options_unique = df_mi_org_gos_prod_options[option_col_dict_gos_option].unique()
            dict_gos_options_unique = dict_lst_gos_prod_options
            # 'name_clean', 'kind', 'i_option', 'option'
            # df_test = semantic_search (df_test, col_name_check, 
            df_test, dict_test_f02_tmp = semantic_search_02 (df_test, col_name_check, 
                        dict_gos_options_unique, df_mi_org_gos_prod_options, 
                        name_col_dict_gos_option, code_col_dict_gos_option, option_col_dict_gos_option,
                        model, 
                        dict_embedding_gos_prod_options_multy, 
                        new_cols_semantic_gos_options, 
                        similarity_threshold, max_sim_entries, n_rows = n_rows) #, debug= debug
            display(df_test.head(2))
            for k,v in dict_test_f02.items():
                try: #if dict_test_f02.get(k) is not None:
                    dict_test_f02[k].update(dict_test_f02_tmp.get(k))
                #else:
                except Exception as err:
                    print(err, "dict_test_f02.get(k) is None", k)
                    print("dict_test_f02.get(k):" , dict_test_f02.get(k))
                    print("dict_test_f02:" , dict_test_f02)
                    print("dict_test_f02_tmp:" , dict_test_f02_tmp)
        else: 
            new_cols_semantic_gos_options = []
      # dict_embedding_gos_prod_options_multy
    else:
        new_cols_semantic_gos = []
        new_cols_semantic_national = []
        new_cols_semantic_gos_options = []
        
    fn_main = fn_check_file.split('.xlsx')[0] + '_processed_pivot'
    
    # print("len(dict_test_f02) - before save:" , len(dict_test_f02))
    df_test_f02 = dict_2_df(dict_test_f02, col_name_check, new_cols_fuzzy, new_cols_semantic,
             new_cols_semantic_gos, new_cols_semantic_national, new_cols_semantic_gos_options)
    # print("df_test_f02.shape - before save:", df_test_f02.shape)
    #fn_save = save_df_to_excel(df_test, data_processed_dir, fn_main)
    # fn_save = save_df_to_excel(df_test_f02.set_index([col_name_check] + new_cols_fuzzy + new_cols_semantic\
    #             + new_cols_semantic_gos + new_cols_semantic_national + new_cols_semantic_gos_options), data_processed_dir, fn_main)
    fn_save = save_df_to_excel(df_test_f02.set_index(list(df_test_f02.columns)[:-1]), data_processed_dir, fn_main, index = True)
    # print("df_test_f02.shape - after save_df_to_excel:", df_test_f02.shape)
    format_excel_cols(data_processed_dir, fn_save, format_cols)
    fn_save_stat = save_stat(df_test, data_processed_dir, fn_check_file, max_sim_entries, similarity_threshold)

    fn_main = fn_check_file.split('.xlsx')[0] + '_processed_flat'
    fn_save = save_df_to_excel(df_test_f02, data_processed_dir, fn_main)
    format_excel_cols(data_processed_dir, fn_save, format_cols)
    
    # work
    # return df_test, df_test_f02
    # debug
    return df_test, df_test_f02, dict_test_f02

def format_excel_cols(data_processed_dir, fn_xls, format_cols):
    wb = load_workbook(os.path.join(data_processed_dir, fn_xls))
    ws = wb.active
    l_alignment=Alignment(horizontal='left', vertical= 'top', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
    r_alignment=Alignment(horizontal='right', vertical= 'top', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
    border = Border( left=Side(border_style="thin", color='FF000000'),
                 right=Side(border_style="thin", color='FF000000'),
                 top=Side(border_style="thin", color='FF000000'),
                 bottom=Side(border_style="thin", color='FF000000'),
     #            diagonal=Side(border_style=None,
     #                          color='FF000000'),
     #            diagonal_direction=0,
     #            outline=Side(border_style="thin", color='FF000000'),
     #            vertical=Side(border_style=None,
     #                          color='FF000000'),
     #            horizontal=Side(border_style=None,
     #                           color='FF000000')
     )
    
    
    # ws.filterMode = True
    last_cell = ws.cell(row=1, column=len(format_cols)) 
    full_range = "A1:" + last_cell.column_letter + str(ws.max_row)
    ws.auto_filter.ref = full_range
    ws.freeze_panes = ws['B2']
    for ic, col_width in enumerate(format_cols):
        cell = ws.cell(row=1, column=ic+1)
        cell.alignment = l_alignment
        ws.column_dimensions[cell.column_letter].width = col_width
    ft = cell.font
    ft = Font(bold=False)
    for row in ws[full_range]: #[1:]
        for cell in row:
            cell.font = ft    
            cell.alignment = l_alignment
            cell.border = border
            
            
        
    wb.save(os.path.join(data_processed_dir, fn_xls))
